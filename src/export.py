"""Export layer for Pokédex Data Pipeline.

Implements Export MVP:
- Writes `data/export/pokedata.xlsx`
- Exports only the `Pokemon` and `Meta` sheets

Rules:
- Sheet structure and column order are binding per spec/export.spec.md
- No derived logic (read transformed data only)
- Export must not modify transformed data
"""

from __future__ import annotations

import os
import csv
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook

from .cache.io import ensure_dir, read_json
from .transform import load_config


DERIVED_DIR = "data/derived"
EXPORT_DIR = "data/export"
EXPORT_PATH = os.path.join(EXPORT_DIR, "pokedata.xlsx")
RAW_ITEM_DIR = "data/raw/item"


def _read_assets_blocks(path: str) -> list[tuple[str, list[tuple[str, str]]]]:
    """Parse config/assets.csv into ordered (heading, rows) blocks.

    Input format:
    - Headings are lines ending with ':' (e.g. 'IMAGES:')
    - Data lines are 'name;address'
    - Blank lines are ignored
    """
    if not os.path.exists(path):
        raise RuntimeError(f"Missing assets config: {path}")

    blocks: list[tuple[str, list[tuple[str, str]]]] = []
    current_heading: Optional[str] = None
    current_rows: list[tuple[str, str]] = []

    with open(path, "r", encoding="utf-8") as f:
        for raw_line in f.read().splitlines():
            line = raw_line.strip()
            if not line:
                continue

            if line.endswith(":"):
                if current_heading is not None:
                    blocks.append((current_heading, current_rows))
                current_heading = line
                current_rows = []
                continue

            if ";" not in line:
                raise RuntimeError(
                    f"Invalid assets.csv line (expected 'name;address'): {raw_line}"
                )
            name, address = line.split(";", 1)
            name = name.strip()
            address = address.strip()
            if not name or not address:
                raise RuntimeError(
                    f"Invalid assets.csv line (empty name/address): {raw_line}"
                )
            if current_heading is None:
                raise RuntimeError(
                    "assets.csv must start with a heading line ending in ':'"
                )
            current_rows.append((name, address))

    if current_heading is not None:
        blocks.append((current_heading, current_rows))

    return blocks


def _read_games_map(path: str) -> Optional[dict[str, str]]:
    """Read config/gamesMap.csv mapping slug -> label.

    Fallback behavior is handled by the caller:
    - If this file is missing/unreadable, return None.

    Expected content: 2 columns (slug, label), with or without a header.
    Delimiter may be ',' or ';' (auto-detected).
    """
    if not os.path.exists(path):
        return None

    try:
        with open(path, "r", encoding="utf-8") as f:
            raw = f.read().splitlines()
    except (OSError, UnicodeError):
        return None

    lines = [ln for ln in (ln.strip() for ln in raw) if ln]
    if not lines:
        return {}

    sample = "\n".join(lines[:5])
    try:
        candidate_delimiters = ",;\t"
        dialect = csv.Sniffer().sniff(sample, delimiters=candidate_delimiters)
    except csv.Error:
        dialect = csv.get_dialect("excel")
        # Best-effort: prefer ';' if present in the first non-empty line.
        if ";" in lines[0] and "," not in lines[0]:
            dialect.delimiter = ";"  # type: ignore[attr-defined]

    mapping: dict[str, str] = {}
    try:
        reader = csv.reader(lines, dialect)
        for row_index, row in enumerate(reader):
            if not row:
                continue

            # Support BOM and allow extra columns; take first two.
            slug = (row[0] or "").strip().lstrip("\ufeff")
            label = (row[1] or "").strip() if len(row) > 1 else ""

            # Skip a header row if present.
            if row_index == 0 and slug.lower() in {"slug", "id", "version_group"}:
                continue
            if row_index == 0 and label.lower() in {"label", "name", "display_name"}:
                continue

            if not slug:
                continue
            mapping[slug] = label
    except csv.Error:
        return None

    return mapping


def _read_derived(path: str) -> Dict[str, Any]:
    data = read_json(path)
    if not data:
        raise RuntimeError(f"Missing or invalid derived file: {path}")
    return data


def _write_row(ws: Any, values: List[Any]) -> None:
    ws.append(values)


def _none_last(value: Any) -> tuple[bool, Any]:
    """Return a deterministic sort key that places None after real values."""
    return (value is None, value)


def _record_str_field(record: Dict[str, Any], field: str) -> str:
    """Best-effort string accessor for sorting dictionaries by string keys."""
    value = record.get(field)
    return value if isinstance(value, str) else ""


def _load_types_order() -> tuple[list[str], bool, str]:
    """Load TypeChart ordering controls from config/typesOrder.json.

    Supports keys used in spec/export.spec.md:
    - typesOrder: list[str]
    - IgnoreRest / ignoreRest: bool
    - IgnoreRestFalseMode / ignoreRestFalseMode: str
    """
    cfg_path = os.path.join("config", "typesOrder.json")
    cfg = read_json(cfg_path) or {}
    if not isinstance(cfg, dict):
        return ([], False, "Put at End")

    order = cfg.get("typesOrder")
    if not isinstance(order, list):
        order = []
    order = [v for v in order if isinstance(v, str) and v.strip()]

    ignore_rest_raw = cfg.get("IgnoreRest")
    if ignore_rest_raw is None:
        ignore_rest_raw = cfg.get("ignoreRest")
    ignore_rest = bool(ignore_rest_raw) if isinstance(ignore_rest_raw, bool) else False

    mode = cfg.get("IgnoreRestFalseMode")
    if mode is None:
        mode = cfg.get("ignoreRestFalseMode")
    mode = mode if isinstance(mode, str) and mode.strip() else "Put at End"

    return (order, ignore_rest, mode)

def _moveset_map(
    learnset_entries: List[Dict[str, Any]],
    version_groups: List[str],
) -> Dict[str, Dict[str, str]]:
    """Build {form_key: {version_group: 'a;b;c'}} mapping for Pokemon sheet."""
    by_form_vg: Dict[str, Dict[str, set[str]]] = {}
    vg_allow = set(version_groups)
    for rec in learnset_entries:
        form_key = rec.get("form_key")
        vg = rec.get("version_group")
        move_key = rec.get("move_key")
        if not isinstance(form_key, str) or not isinstance(vg, str) or \
           not isinstance(move_key, str):
            continue
        if vg not in vg_allow:
            continue
        by_form_vg.setdefault(form_key, {}).setdefault(vg, set()).add(move_key)

    out: Dict[str, Dict[str, str]] = {}
    for form_key, vg_map in by_form_vg.items():
        out[form_key] = {}
        for vg, moves in vg_map.items():
            out[form_key][vg] = ";".join(sorted(moves))
    return out


def _normalize_version_groups(raw: Any) -> List[Dict[str, Any]]:
    entries: List[Dict[str, Any]] = []
    if not isinstance(raw, list):
        return entries
    for item in raw:
        if isinstance(item, str) and item.strip():
            entries.append({"id": item, "label": item, "versions": []})
            continue
        if not isinstance(item, dict):
            continue
        vg_id = item.get("id") or item.get("slug") or item.get("key")
        if not isinstance(vg_id, str) or not vg_id.strip():
            continue
        label = (
            item.get("name")
            or item.get("label")
            or item.get("display_name")
            or vg_id
        )
        if not isinstance(label, str) or not label.strip():
            label = vg_id
        versions_raw = item.get("versions")
        versions: List[str] = []
        if isinstance(versions_raw, list):
            versions = [v for v in versions_raw if isinstance(v, str) and v.strip()]
        entries.append({"id": vg_id, "label": label, "versions": versions})
    return entries


def _sorted_display_values(values: Iterable[Any]) -> List[str]:
    unique: set[str] = set()
    for value in values:
        if isinstance(value, str):
            trimmed = value.strip()
            if trimmed:
                unique.add(trimmed)
    return sorted(unique, key=lambda v: (v.lower(), v))


def _display_names_from_keys(
    keys: Iterable[str], display_map: Dict[str, Optional[str]]
) -> List[str]:
    return _sorted_display_values(display_map.get(key) for key in keys)


def _load_item_version_membership(version_groups: List[str]) -> Dict[str, set[str]]:
    membership: Dict[str, set[str]] = {vg: set() for vg in version_groups}
    if not os.path.isdir(RAW_ITEM_DIR):
        return membership

    vg_allow = set(version_groups)
    for name in os.listdir(RAW_ITEM_DIR):
        if not name.lower().endswith(".json"):
            continue
        path = os.path.join(RAW_ITEM_DIR, name)
        raw = read_json(path) or {}
        data = raw.get("data") or {}
        if not isinstance(data, dict):
            continue
        item_key = data.get("name")
        if not isinstance(item_key, str):
            continue

        flavor_entries = data.get("flavor_text_entries")
        if not isinstance(flavor_entries, list):
            continue

        seen_vgs: set[str] = set()
        for entry in flavor_entries:
            if not isinstance(entry, dict):
                continue
            vg_name = ((entry.get("version_group") or {}).get("name"))
            if isinstance(vg_name, str) and vg_name in vg_allow:
                seen_vgs.add(vg_name)

        for vg in seen_vgs:
            membership.setdefault(vg, set()).add(item_key)

    return membership


def _build_gameversion_columns(
    *,
    version_groups: List[str],
    learnset_entries: List[Dict[str, Any]],
    pokemon_forms: List[Dict[str, Any]],
    moves: List[Dict[str, Any]],
    abilities: List[Dict[str, Any]],
    items: List[Dict[str, Any]],
    item_membership: Dict[str, set[str]],
) -> List[Tuple[str, List[str]]]:
    vg_to_forms: Dict[str, set[str]] = {vg: set() for vg in version_groups}
    vg_to_moves: Dict[str, set[str]] = {vg: set() for vg in version_groups}

    for entry in learnset_entries:
        form_key = entry.get("form_key")
        move_key = entry.get("move_key")
        vg = entry.get("version_group")
        if not (
            isinstance(form_key, str)
            and isinstance(move_key, str)
            and isinstance(vg, str)
        ):
            continue
        if vg not in vg_to_forms:
            continue
        vg_to_forms[vg].add(form_key)
        vg_to_moves[vg].add(move_key)

    form_display: Dict[str, Optional[str]] = {}
    for rec in pokemon_forms:
        if not isinstance(rec, dict):
            continue
        form_key = rec.get("form_key")
        if not isinstance(form_key, str):
            continue
        display_name = rec.get("display_name")
        form_display[form_key] = display_name if isinstance(display_name, str) else None

    move_display: Dict[str, Optional[str]] = {}
    for rec in moves:
        if not isinstance(rec, dict):
            continue
        move_key = rec.get("move_key")
        if not isinstance(move_key, str):
            continue
        display_name = rec.get("display_name")
        move_display[move_key] = display_name if isinstance(display_name, str) else None

    ability_display: Dict[str, Optional[str]] = {}
    for rec in abilities:
        if not isinstance(rec, dict):
            continue
        ability_key = rec.get("ability_key")
        if not isinstance(ability_key, str):
            continue
        display_name = rec.get("display_name")
        ability_display[ability_key] = (
            display_name if isinstance(display_name, str) else None
        )

    ability_display_to_key: Dict[str, str] = {}
    ability_display_to_key_ci: Dict[str, str] = {}
    for rec in abilities:
        if not isinstance(rec, dict):
            continue
        ability_key = rec.get("ability_key")
        display_name = rec.get("display_name")
        if not (isinstance(ability_key, str) and isinstance(display_name, str)):
            continue
        trimmed = display_name.strip()
        if not trimmed:
            continue
        ability_display_to_key[trimmed] = ability_key
        ability_display_to_key_ci[trimmed.lower()] = ability_key

    item_display: Dict[str, Optional[str]] = {}
    for rec in items:
        if not isinstance(rec, dict):
            continue
        item_key = rec.get("item_key")
        if not isinstance(item_key, str):
            continue
        display_name = rec.get("display_name")
        item_display[item_key] = display_name if isinstance(display_name, str) else None

    form_abilities: Dict[str, List[str]] = {}
    for rec in pokemon_forms:
        if not isinstance(rec, dict):
            continue
        form_key = rec.get("form_key")
        if not isinstance(form_key, str):
            continue
        resolved_ability_keys: List[str] = []
        for field in ("ability1", "ability2", "hidden_ability"):
            value = rec.get(field)
            if not isinstance(value, str):
                continue
            trimmed = value.strip()
            if not trimmed:
                continue
            ability_key = ability_display_to_key.get(trimmed)
            if not ability_key:
                ability_key = ability_display_to_key_ci.get(trimmed.lower())
            if ability_key:
                resolved_ability_keys.append(ability_key)
        form_abilities[form_key] = resolved_ability_keys

    vg_to_abilities: Dict[str, set[str]] = {vg: set() for vg in version_groups}
    for vg, form_keys in vg_to_forms.items():
        vg_ability_keys: set[str] = set()
        for form_key in form_keys:
            resolved = form_abilities.get(form_key)
            if not resolved:
                continue
            vg_ability_keys.update(resolved)
        vg_to_abilities[vg] = vg_ability_keys

    all_form_keys: set[str] = set()
    all_move_keys: set[str] = set()
    all_ability_keys: set[str] = set()
    all_item_keys: set[str] = set()

    for keys in vg_to_forms.values():
        all_form_keys.update(keys)
    for keys in vg_to_moves.values():
        all_move_keys.update(keys)
    for keys in vg_to_abilities.values():
        all_ability_keys.update(keys)
    for keys in item_membership.values():
        all_item_keys.update(keys)

    columns: List[Tuple[str, List[str]]] = [
        ("MOVES_ALL", _display_names_from_keys(all_move_keys, move_display)),
        ("POKEMON_ALL", _display_names_from_keys(all_form_keys, form_display)),
        ("ABILITIES_ALL", _display_names_from_keys(all_ability_keys, ability_display)),
        ("ITEMS_ALL", _display_names_from_keys(all_item_keys, item_display)),
    ]

    for vg in version_groups:
        columns.append((f"MOVES_{vg}", _display_names_from_keys(vg_to_moves[vg], move_display)))
        columns.append((f"POKEMON_{vg}", _display_names_from_keys(vg_to_forms[vg], form_display)))
        columns.append(
            (
                f"ABILITIES_{vg}",
                _display_names_from_keys(vg_to_abilities[vg], ability_display),
            )
        )
        columns.append(
            (
                f"ITEMS_{vg}",
                _display_names_from_keys(item_membership.get(vg, set()), item_display),
            )
        )

    return columns

def run_export_mvp(config_path: str = "config/config.json") -> int:
    """Export MVP: create workbook with Pokemon + Meta sheets."""
    cfg = load_config(config_path)
    vg_entries = _normalize_version_groups(cfg.get("version_groups", []))
    version_groups = [entry["id"] for entry in vg_entries]

    forms_path = os.path.join(DERIVED_DIR, "pokemon_forms.json")
    meta_path = os.path.join(DERIVED_DIR, "meta.json")

    forms_doc = _read_derived(forms_path)
    meta_doc = _read_derived(meta_path)

    forms = forms_doc.get("pokemon_forms")
    if not isinstance(forms, list):
        raise RuntimeError("Derived pokemon_forms.json missing pokemon_forms list")

    ensure_dir(EXPORT_DIR)

    wb = Workbook()
    # Remove default sheet so we control sheet order.
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)

    # Sheet: Pokemon
    ws_pokemon = wb.create_sheet("Pokemon")

    headers: List[str] = [
        "DEX_ID",
        "FORM_KEY",
        "DISPLAY_NAME",
        "FORM_GROUP",
        "TYPE1",
        "TYPE2",
        "HP",
        "ATK",
        "DEF",
        "SPA",
        "SPD",
        "SPE",
        "TOTAL",
        "ABILITY1",
        "ABILITY2",
        "HIDDEN_ABILITY",
        "HEIGHT_M",
        "WEIGHT_KG",
        "ABOUT",
        "SPRITE",
        "SHINY_SPRITE",
    ]
    for vg in version_groups:
        headers.append(f"MOVESET_{vg}")
    _write_row(ws_pokemon, headers)

    for rec in forms:
        if not isinstance(rec, dict):
            continue
        row: List[Any] = [
            rec.get("dex_id"),
            rec.get("form_key"),
            rec.get("display_name"),
            rec.get("form_group"),
            rec.get("type1"),
            rec.get("type2"),
            rec.get("base_hp"),
            rec.get("base_atk"),
            rec.get("base_def"),
            rec.get("base_spa"),
            rec.get("base_spd"),
            rec.get("base_spe"),
            rec.get("base_total"),
            rec.get("ability1"),
            rec.get("ability2"),
            rec.get("hidden_ability"),
            rec.get("height_m"),
            rec.get("weight_kg"),
            rec.get("about"),
            rec.get("sprite_url"),
            rec.get("shiny_sprite_url"),
        ]
        # MVP: do not compute movesets; emit empty cells.
        row.extend([None] * len(version_groups))
        _write_row(ws_pokemon, row)

    # Sheet: Meta
    ws_meta = wb.create_sheet("Meta")
    _write_row(ws_meta, ["KEY", "VALUE"])

    def meta_val(key: str) -> Optional[Any]:
        return meta_doc.get(key)

    required_meta_rows: List[List[Any]] = [
        ["generated_at", meta_val("generated_at")],
        ["source", meta_val("source")],
        ["pokeapi_base_url", cfg.get("pokeapi_base_url")],
        ["about_language", meta_val("about_language")],
        [
            "version_groups",
            ",".join(meta_val("version_groups") or [])
            if isinstance(meta_val("version_groups"), list)
            else "",
        ],
        ["pipeline_version", meta_val("pipeline_version")],
    ]
    for r in required_meta_rows:
        _write_row(ws_meta, r)

    wb.save(EXPORT_PATH)
    print(f"export mvp: wrote {EXPORT_PATH}")
    return 0


def run_export_extended(config_path: str = "config/config.json") -> int:
    """Export Extended: add Learnsets/Moves/Items/Abilities/Natures/Evolutions/TypeChart/Assets."""
    cfg = load_config(config_path)
    vg_entries = _normalize_version_groups(cfg.get("version_groups", []))
    version_groups = [entry["id"] for entry in vg_entries]

    forms_doc = _read_derived(os.path.join(DERIVED_DIR, "pokemon_forms.json"))
    learnsets_doc = _read_derived(os.path.join(DERIVED_DIR, "learnset_entries.json"))
    moves_doc = _read_derived(os.path.join(DERIVED_DIR, "moves.json"))
    items_doc = _read_derived(os.path.join(DERIVED_DIR, "items.json"))
    abilities_doc = _read_derived(os.path.join(DERIVED_DIR, "abilities.json"))
    natures_doc = _read_derived(os.path.join(DERIVED_DIR, "natures.json"))
    evolutions_doc = _read_derived(os.path.join(DERIVED_DIR, "evolution_edges.json"))
    type_chart_doc = _read_derived(os.path.join(DERIVED_DIR, "type_chart.json"))
    types_doc = _read_derived(os.path.join(DERIVED_DIR, "types.json"))
    meta_doc = _read_derived(os.path.join(DERIVED_DIR, "meta.json"))

    forms = forms_doc.get("pokemon_forms")
    learnset_entries = learnsets_doc.get("learnset_entries")
    moves = moves_doc.get("moves")
    items = items_doc.get("items")
    abilities = abilities_doc.get("abilities")
    natures = natures_doc.get("natures")
    evolutions = evolutions_doc.get("evolution_edges")
    type_keys = type_chart_doc.get("type_keys")
    type_matrix = type_chart_doc.get("matrix")
    types = types_doc.get("types")

    if not isinstance(forms, list):
        raise RuntimeError("pokemon_forms.json missing pokemon_forms list")
    if not isinstance(learnset_entries, list):
        raise RuntimeError("learnset_entries.json missing learnset_entries list")
    if not isinstance(moves, list):
        raise RuntimeError("moves.json missing moves list")
    if not isinstance(items, list):
        raise RuntimeError("items.json missing items list")
    if not isinstance(abilities, list):
        raise RuntimeError("abilities.json missing abilities list")
    if not isinstance(natures, list):
        raise RuntimeError("natures.json missing natures list")
    if not isinstance(evolutions, list):
        raise RuntimeError("evolution_edges.json missing evolution_edges list")
    if not isinstance(type_keys, list) or not all(isinstance(t, str) for t in type_keys):
        raise RuntimeError("type_chart.json missing type_keys list")
    if not isinstance(type_matrix, list):
        raise RuntimeError("type_chart.json missing matrix")
    if not isinstance(types, list):
        raise RuntimeError("types.json missing types list")

    ensure_dir(EXPORT_DIR)

    wb = Workbook()
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)

    # Sheet: Pokemon
    ws_pokemon = wb.create_sheet("Pokemon")
    pokemon_headers: List[str] = [
        "DEX_ID",
        "FORM_KEY",
        "DISPLAY_NAME",
        "FORM_GROUP",
        "TYPE1",
        "TYPE2",
        "HP",
        "ATK",
        "DEF",
        "SPA",
        "SPD",
        "SPE",
        "TOTAL",
        "ABILITY1",
        "ABILITY2",
        "HIDDEN_ABILITY",
        "HEIGHT_M",
        "WEIGHT_KG",
        "ABOUT",
        "SPRITE",
        "SHINY_SPRITE",
    ]
    for vg in version_groups:
        pokemon_headers.append(f"MOVESET_{vg}")
    _write_row(ws_pokemon, pokemon_headers)

    movesets = _moveset_map(learnset_entries, version_groups)

    # Row ordering: DEX_ID, FORM_KEY
    forms_sorted: List[Dict[str, Any]] = [rec for rec in forms if isinstance(rec, dict)]
    forms_sorted.sort(key=lambda r: (r.get("dex_id"), r.get("form_key")))
    for rec in forms_sorted:
        form_key = rec.get("form_key")
        row: List[Any] = [
            rec.get("dex_id"),
            form_key,
            rec.get("display_name"),
            rec.get("form_group"),
            rec.get("type1"),
            rec.get("type2"),
            rec.get("base_hp"),
            rec.get("base_atk"),
            rec.get("base_def"),
            rec.get("base_spa"),
            rec.get("base_spd"),
            rec.get("base_spe"),
            rec.get("base_total"),
            rec.get("ability1"),
            rec.get("ability2"),
            rec.get("hidden_ability"),
            rec.get("height_m"),
            rec.get("weight_kg"),
            rec.get("about"),
            rec.get("sprite_url"),
            rec.get("shiny_sprite_url"),
        ]
        if isinstance(form_key, str):
            for vg in version_groups:
                row.append(movesets.get(form_key, {}).get(vg))
        else:
            row.extend([None] * len(version_groups))
        _write_row(ws_pokemon, row)

    # Sheet: Learnsets
    ws_ls = wb.create_sheet("Learnsets")
    _write_row(
        ws_ls,
        ["FORM_KEY", "DISPLAY_NAME", "VERSION_GROUP", "MOVE_KEY", "METHOD", "LEVEL"],
    )
    ls_sorted: List[Dict[str, Any]] = [
        rec for rec in learnset_entries if isinstance(rec, dict)
    ]
    ls_sorted.sort(
        key=lambda r: (
            r.get("form_key"),
            r.get("version_group"),
            r.get("move_key"),
            r.get("method"),
            _none_last(r.get("level")),
        )
    )
    for rec in ls_sorted:
        _write_row(
            ws_ls,
            [
                rec.get("form_key"),
                rec.get("display_name"),
                rec.get("version_group"),
                rec.get("move_key"),
                rec.get("method"),
                rec.get("level"),
            ],
        )

    # Sheet: Moves
    ws_moves = wb.create_sheet("Moves")
    _write_row(
        ws_moves,
        [
            "MOVE_KEY",
            "DISPLAY_NAME",
            "TYPE",
            "CATEGORY",
            "POWER",
            "ACCURACY",
            "PP",
            "PRIORITY",
            "EFFECT_SHORT",
        ],
    )
    moves_sorted: List[Dict[str, Any]] = [rec for rec in moves if isinstance(rec, dict)]
    moves_sorted.sort(key=lambda r: _record_str_field(r, "move_key"))
    for rec in moves_sorted:
        _write_row(
            ws_moves,
            [
                rec.get("move_key"),
                rec.get("display_name"),
                rec.get("type"),
                rec.get("category"),
                rec.get("power"),
                rec.get("accuracy"),
                rec.get("pp"),
                rec.get("priority"),
                rec.get("effect_short"),
            ],
        )

    # Sheet: Items
    ws_items = wb.create_sheet("Items")
    _write_row(ws_items, ["ITEM_KEY", "DISPLAY_NAME", "CATEGORY", "EFFECT_SHORT"])
    items_sorted: List[Dict[str, Any]] = [rec for rec in items if isinstance(rec, dict)]
    items_sorted.sort(key=lambda r: _record_str_field(r, "item_key"))
    for rec in items_sorted:
        _write_row(
            ws_items,
            [
                rec.get("item_key"),
                rec.get("display_name"),
                rec.get("category"),
                rec.get("effect_short"),
            ],
        )

    # Sheet: Abilities
    ws_abilities = wb.create_sheet("Abilities")
    _write_row(ws_abilities, ["ABILITY_KEY", "DISPLAY_NAME", "EFFECT_SHORT"])
    abilities_sorted: List[Dict[str, Any]] = [
        rec for rec in abilities if isinstance(rec, dict)
    ]
    abilities_sorted.sort(key=lambda r: _record_str_field(r, "ability_key"))
    for rec in abilities_sorted:
        _write_row(
            ws_abilities,
            [rec.get("ability_key"), rec.get("display_name"), rec.get("effect_short")],
        )

    # Sheet: Natures
    ws_natures = wb.create_sheet("Natures")
    _write_row(
        ws_natures,
        ["NATURE_KEY", "DISPLAY_NAME", "INCREASED_STAT", "DECREASED_STAT"],
    )
    natures_sorted: List[Dict[str, Any]] = [
        rec for rec in natures if isinstance(rec, dict)
    ]
    natures_sorted.sort(key=lambda r: _record_str_field(r, "nature_key"))
    for rec in natures_sorted:
        _write_row(
            ws_natures,
            [
                rec.get("nature_key"),
                rec.get("display_name"),
                rec.get("increased_stat"),
                rec.get("decreased_stat"),
            ],
        )

    types_sorted: List[Dict[str, Any]] = [rec for rec in types if isinstance(rec, dict)]
    types_sorted.sort(key=lambda r: _record_str_field(r, "type_key"))

    # Sheet: Evolutions
    # Note: data-contract.md defines FROM_DEX_ID/TO_DEX_ID etc; export writes derived values.
    ws_evo = wb.create_sheet("Evolutions")
    _write_row(
        ws_evo,
        [
            "FROM_DEX_ID",
            "TO_DEX_ID",
            "TRIGGER",
            "MIN_LEVEL",
            "ITEM_KEY",
            "TIME_OF_DAY",
            "MIN_HAPPINESS",
            "KNOWN_MOVE_KEY",
            "KNOWN_MOVE_TYPE",
            "LOCATION",
            "GENDER",
            "HELD_ITEM_KEY",
        ],
    )
    evo_sorted: List[Dict[str, Any]] = [
        rec for rec in evolutions if isinstance(rec, dict)
    ]
    evo_sorted.sort(
        key=lambda r: (
            _none_last(r.get("from_dex_id")),
            _none_last(r.get("to_dex_id")),
            _none_last(r.get("trigger")),
            _none_last(r.get("min_level")),
            _none_last(r.get("item_key")),
        )
    )
    for rec in evo_sorted:
        _write_row(
            ws_evo,
            [
                rec.get("from_dex_id"),
                rec.get("to_dex_id"),
                rec.get("trigger"),
                rec.get("min_level"),
                rec.get("item_key"),
                rec.get("time_of_day"),
                rec.get("min_happiness"),
                rec.get("known_move_key"),
                rec.get("known_move_type"),
                rec.get("location"),
                rec.get("gender"),
                rec.get("held_item_key"),
            ],
        )

    # Sheet: TypeChart (matrix)
    # Matrix layout: attacking type (rows) x defending type (columns).
    # Ordering for both axes is controlled by config/typesOrder.json.
    ws_tc = wb.create_sheet("TypeChart")

    type_key_to_display: Dict[str, str] = {}
    display_to_type_key: Dict[str, str] = {}
    for rec in types_sorted:
        tk = rec.get("type_key")
        dn = rec.get("display_name")
        if isinstance(tk, str) and isinstance(dn, str):
            type_key_to_display[tk] = dn
            display_to_type_key[dn] = tk

    type_key_to_index = {
        tk: idx for idx, tk in enumerate(type_keys) if isinstance(tk, str)
    }

    order_list, ignore_rest, ignore_rest_mode = _load_types_order()
    final_types: list[str] = [dn for dn in order_list if dn in display_to_type_key]

    if not ignore_rest:
        rest = sorted([dn for dn in display_to_type_key if dn not in set(final_types)])
        # Only defined mode currently in config is "Put at End".
        if isinstance(ignore_rest_mode, str) and ignore_rest_mode.strip().lower() == "put at end":
            final_types.extend(rest)
        else:
            final_types.extend(rest)

    header = ["ATTACKING_TYPE"] + list(final_types)
    _write_row(ws_tc, header)

    for atk_dn in final_types:
        atk_key = display_to_type_key.get(atk_dn)
        atk_idx = type_key_to_index.get(atk_key) if isinstance(atk_key, str) else None
        if atk_idx is None:
            continue

        src_row = type_matrix[atk_idx] if atk_idx < len(type_matrix) else []
        out_row: List[Any] = [atk_dn]

        for def_dn in final_types:
            def_key = display_to_type_key.get(def_dn)
            def_idx = type_key_to_index.get(def_key) if isinstance(def_key, str) else None
            if def_idx is None:
                out_row.append(None)
                continue

            multiplier = (
                src_row[def_idx]
                if isinstance(src_row, list) and def_idx < len(src_row)
                else None
            )
            out_row.append(multiplier)

        _write_row(ws_tc, out_row)

    # Sheet: Types
    ws_types = wb.create_sheet("Types")
    _write_row(ws_types, ["DISPLAY_NAME", "TYPE_KEY", "ICON_URL"])
    for rec in types_sorted:
        _write_row(ws_types, [rec.get("display_name"), rec.get("type_key"), rec.get("icon_url")])

    # Sheet: Meta
    ws_meta = wb.create_sheet("Meta")
    _write_row(ws_meta, ["KEY", "VALUE"])
    required_meta_rows: List[List[Any]] = [
        ["generated_at", meta_doc.get("generated_at")],
        ["source", meta_doc.get("source")],
        ["pokeapi_base_url", cfg.get("pokeapi_base_url")],
        ["about_language", meta_doc.get("about_language")],
        [
            "version_groups",
            ",".join(meta_doc.get("version_groups") or [])
            if isinstance(meta_doc.get("version_groups"), list)
            else "",
        ],
        ["pipeline_version", meta_doc.get("pipeline_version")],
    ]
    for r in required_meta_rows:
        _write_row(ws_meta, r)

    # Sheet: Assets
    # Layout per spec/export.spec.md §4.10:
    # - each heading creates a new 2-column block (name, address), starting at row 1
    # - at the end add a new 1-column block "GAMES" listing configured version group labels
    ws_assets = wb.create_sheet("Assets")
    assets_blocks = _read_assets_blocks(os.path.join("config", "assets.csv"))
    for block_index, (heading, rows) in enumerate(assets_blocks):
        start_col = 1 + (block_index * 2)
        ws_assets.cell(row=1, column=start_col, value=heading)
        for i, (name, address) in enumerate(rows):
            ws_assets.cell(row=2 + i, column=start_col, value=name)
            ws_assets.cell(row=2 + i, column=start_col + 1, value=address)

    # Append final block: GAMES
    games_block_col = 1 + (len(assets_blocks) * 2)
    ws_assets.cell(row=1, column=games_block_col, value="GAMES")

    games_map_path = os.path.join("config", "gamesMap.csv")
    games_map = _read_games_map(games_map_path)

    # Source of truth (IDs): config/config.json -> version_groups (preserve order)
    for row_offset, entry in enumerate(vg_entries):
        vg_id = entry.get("id")
        if not isinstance(vg_id, str) or not vg_id.strip():
            continue
        fallback_label = entry.get("label") if isinstance(entry.get("label"), str) else vg_id
        if not fallback_label or not fallback_label.strip():
            fallback_label = vg_id
        label = fallback_label
        if isinstance(games_map, dict):
            mapped = games_map.get(vg_id)
            if isinstance(mapped, str) and mapped.strip():
                label = mapped
        ws_assets.cell(row=2 + row_offset, column=games_block_col, value=label)

    ws_assets.cell(row=2 + len(vg_entries), column=games_block_col, value="All")

    # Sheet: GAMEVERSIONS
    item_membership = _load_item_version_membership(version_groups)
    gameversion_columns = _build_gameversion_columns(
        version_groups=version_groups,
        learnset_entries=[rec for rec in learnset_entries if isinstance(rec, dict)],
        pokemon_forms=[rec for rec in forms if isinstance(rec, dict)],
        moves=[rec for rec in moves if isinstance(rec, dict)],
        abilities=[rec for rec in abilities if isinstance(rec, dict)],
        items=[rec for rec in items if isinstance(rec, dict)],
        item_membership=item_membership,
    )
    ws_gameversions = wb.create_sheet("GAMEVERSIONS")
    for col_index, (header, values) in enumerate(gameversion_columns, start=1):
        ws_gameversions.cell(row=1, column=col_index, value=header)
        for row_index, value in enumerate(values, start=2):
            ws_gameversions.cell(row=row_index, column=col_index, value=value)

    wb.save(EXPORT_PATH)
    print(f"export extended: wrote {EXPORT_PATH}")
    return 0


def run_export_production(config_path: str = "config/config.json") -> int:
    """Export Production: full workbook, fail fast on missing inputs."""
    # Current implementation exports the full workbook (all sheets) already.
    # Production mode's distinct behavior is to fail fast if required inputs
    # are missing; run_export_extended() already raises on missing derived files.
    rc = run_export_extended(config_path)
    print("export production: ok")
    return rc
